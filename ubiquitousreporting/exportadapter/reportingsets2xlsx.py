# -*- coding: utf-8 -*-
import os
import sys
from sdsct.ubiquitousreporting.dataobjects.report_generic import CodePlan
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabHead
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabReportDef
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabDef
from sdsct.ubiquitousreporting.exportadapter.reportingset2export import ReportingSet2Export
import openpyxl
from openpyxl.workbook import Workbook
import logging

logging.basicConfig(level=logging.INFO)


class ReportingSets2Xlsx(ReportingSet2Export):

    file_extension = ".xlsx"

    def export_tabreport(self, tabreport: TabReportDef, output_name='tabband', testmode=False):
        logging.info('EXPORT tabreport to XLSX')
        wb = Workbook()
        sheet = wb.active
        x, y = (1, 1)
        spacing_between_tables = 3  # cells
        table_offset = 0
        for tabdef_name in tabreport.tabdef_order:
            table_offset += spacing_between_tables + self.plot_tabdef(tabdef=tabreport[tabdef_name],
                                                                      sheet=sheet, x=x, y=y + table_offset)
        wb.save(self.output_directory + os.sep + output_name + self.file_extension)
        logging.info('EXPORT tabreport to XLSX finished')

    def plot_tabdef(self, tabdef: TabDef, sheet, x, y, head: TabHead = None):
        height_of_header = 3
        sheet.cell(row=y, column=x).value = tabdef.title or tabdef.name  ## fixme push to tabdef.
        reportingset = tabdef.reportingset
        tabhead_xlsx = TabHead()
        tabhead_xlsx.append(CodePlan(name='TOTAL', data=['TOTAL']))
        for head in tabdef.table_head:
            if isinstance(head, TabHead):
                for cp in head:
                    tabhead_xlsx.append(cp)
            elif isinstance(head, CodePlan):
                tabhead_xlsx.append(head)

        # plot code labels
        code_order = reportingset.split_main.key_order
        for code_count, code in enumerate(code_order):
            print(tabdef.reportingset.split_main.keys())
            code_label = tabdef.reportingset.split_main[str(int(code))]

            sheet.cell(column=x, row=y + height_of_header + code_count).value = code_label

        # plot headers
        head_offset = x + 1
        for head_count, head in enumerate(tabhead_xlsx):
            sheet.cell(row=y, column=head_offset).value = head.title
            for headcode in head.key_order:
                sheet.cell(row=y + 1, column=head_offset).value = head[headcode]
                for column_reporting in self.columns_reporting:
                    sheet.cell(row=y + 2, column=head_offset).value = column_reporting
                    for code_count, code in enumerate(code_order):
                        dataset_name = headcode if headcode == 'TOTAL' else (head.variables[0] + '_' + str(headcode))
                        dataset_relevant = reportingset[dataset_name]
                        try:
                            value_display = dataset_relevant[column_reporting][code]
                        except KeyError:
                            try:
                                value_display = dataset_relevant[column_reporting][str(code)]
                            except KeyError:
                                value_display = 0
                        sheet.cell(row=y + height_of_header + code_count, column=head_offset).value = value_display
                    head_offset += 1
        return len(code_order) + height_of_header
