# -*- coding: utf-8 -*-
from sdsct.ubiquitousreporting.dataobjects.report_generic import CodePlan
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabHead
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabReportDef
from sdsct.ubiquitousreporting.dataobjects.report_tab import TabDef
from sdsct.ubiquitousreporting.exportadapter.ReportingSet2ExportFormat import ReportingSet2ExportFormat
import openpyxl
from openpyxl.workbook import Workbook
import logging

logging.basicConfig(level=logging.INFO)


class ReportingSet2Xlsx(ReportingSet2ExportFormat):
    # def __init__(self, output_directory):
    #     self.title = "Tabellenbandtitel"
    #     self.subtitle = "Tabellenbanduntertitel"
    #     output_directory = output_directory.replace('\\', '/')  # TODO: put global
    #     self.outputDir = output_directory if output_directory.endswith('/') else output_directory
    #     self.columnType = {'COUNT': 'Anzahl',
    #                        'PCT': 'Prozent',
    #                        'SUM': 'Summe',
    #                        'AVG': 'Durchschnitt',
    #                        'COUNT_W': 'Anzahl gew.',
    #                        'PCT_W': 'Prozent gew.',
    #                        'PCT_GN_W': 'ProzentBasis gew.',
    #                        'SUM_W': 'Summe gew.',
    #                        'AVG_W': 'Durchschnitt gew.'}

    def export_tabreport(self, tabreport: TabReportDef, output_name='tabband', testmode=False):
        logging.info('EXPORT tabreport to XLS')
        wb = Workbook()
        sheet = wb.active
        x, y = (1, 1)
        spacing_between_tables = 2  # cells
        table_offset = 0
        for tabdef_name in tabreport.tabdef_order:
            table_offset += spacing_between_tables + self.plot_tabdef(tabdef=tabreport[tabdef_name],
                                                                      sheet=sheet, x=x, y=y + table_offset)




            #         if tabdef.filter_rs:
            #             if tabdef.filter_rs and isinstance(tabdef.filter_rs, str):
            #                 tabdef.reportingset.content_misc['filterText'] = tabdef.filter_rs
            #             else:
            #                 tabdef.reportingset.content_misc['filterText'] = tabdef.filter_rs.label
            #         if tabdef.section:
            #             tabdef.reportingset.content_misc['section'] = tabdef.section
            #         if tabdef.reporting_columns:
            #             self.reporting_columns = tabdef.reporting_columns

            #         # try:
            #         #     sheet.write(y+tabOffset,0, tabDef.title.decode('utf-8'))
            #         # except UnicodeEncodeError:
            #         #     sheet.write(y+tabOffset,0, tabDef.title.encode('utf-8').decode('utf-8'))
            #
            #         if tabdef.table_head:
            #             i = 0
            #             for head in tabdef.table_head:
            #                 if i == 0:
            #                     if isinstance(head, CodePlan):
            #                         table_offset += self.plot_tabdef(tabdef, head, sheet, x, y + table_offset) + 6
            #                     elif isinstance(head, TabHead) and head.is_subhead is True:
            #                         table_offset += self.plot_tabdef(tabdef, head, sheet, x, y + table_offset) + 6
            #                     else:
            #                         print(type(head))
            #                         raise Exception("WTF")
            #                 i += 1
        wb.save('m:/tabband.xlsx')

    #     if output_name[-4:] == ".xls":  # todo auslagern
    #         wb.save(self.outputDir + '/' +  output_name)
    #     else:
    #         wb.save(self.outputDir + '/' +  output_name + '.xls')
    #     logging.info('Tabreport to XLS exported!')
    #

    def plot_tabdef(self, tabdef: TabDef, sheet, x, y, head: TabHead = None):
        sheet.cell(row=y, column=x).value = tabdef.title or tabdef.name  ## fixme push to tabdef.
        reportingset = tabdef.reportingset
        # tabhead_xls = TabHead()
        # for head in tabdef.table_head:
        #     if isinstance(head, TabHead):
        #         for cp in head:
        #             tabhead_xls.append(cp)
        #     elif isinstance(head, CodePlan):
        #         tabhead_xls.append(head)
        # x, y = (x, y)
        # Plot Code labels
        height_of_header = 3
        code_order = reportingset.split_main.key_order
        for code_count, code in enumerate(code_order):
            # try:
            code_label = tabdef.reportingset.split_main[str(code)]
            # except:
            #     try:
            #         code_label = tabdef.reportingset.split_main[int(code)]
            #     except:
            #         code_label = tabdef.reportingset.split_main[code]
            sheet.cell(column=x, row=y + height_of_header + code_count).value = code_label

        sheet.cell(row=y, column=x + 1).value = 'TOTAL'
        for offset_code, label in enumerate(self.columns_reporting):
            # column labels:
            sheet.cell(column=x + offset_code + 1, row=y + 2).value = label
            for code_count, code in enumerate(code_order):
                try:
                    sheet.cell(row=y + height_of_header + code_count, column=x + offset_code + 1).value = \
                        reportingset['TOTAL'][label][code]
                except KeyError:
                    try:
                        sheet.cell(row=y + height_of_header + code_count, column=x + offset_code + 1).value = \
                            reportingset['TOTAL'][label][str(code)]
                    except KeyError:
                        sheet.cell(row=y + height_of_header + code_count, column=x + offset_code + 1).value = 0

        # x += len(self.reporting_columns)
        return len(code_order) + height_of_header

        #     for head in tabhead_xls:
        #         sheet.write(y, x, head.title)
        #         offset_code = 0
        #         split_sub = rs.splits_sub[head.variables[0]]
        #         for code in split_sub.key_order:
        #             #        relevantDataSetNames = [subHead.variables[0] + "_" + str(code) for subHead in tableHeads for code in subHead.keyOrder]
        #
        #             #        relevantDataSetNames.insert(0, 'TOTAL')
        #             #       dataSets = {name:rs.dataSets[name] for name in relevantDataSetNames}
        #
        #             sheet.write(y + 1, x + offset_code, head[code])
        #             # try:
        #             #     sheet.write(y+1,x+codeOffset, head[code].decode('utf-8'))
        #             # except UnicodeEncodeError:
        #             #     sheet.write(y+1,x+codeOffset, head[code].encode('utf-8').decode('utf-8'))
        #             # try:
        #
        #             # except:
        #             #    sheet.write(y+1,x+codeOffset, code)
        #             for label in self.reporting_columns:
        #                 # sheet.write(y+2,x+codeOffset, label.decode('utf-8'))
        #                 sheet.write(y + 2, x + offset_code, label)
        #                 # codeOffset += len(self.reporting_columns)
        #                 code_count = 0
        #                 for valCode in code_order:
        #                     dataset = rs[head.variables[0] + "_" + str(code)]
        #                     sheet.write(y + 3 + code_count, x + offset_code, dataset.get(label, {}).get(str(valCode), 0))
        #                     code_count += 1
        #                 if label == self.reporting_columns[0]:
        #                     sheet.write(y + 3 + code_count, x + offset_code, dataset.sql_result_str)
        #                 offset_code += 1
        #         x += len(self.reporting_columns) * len(split_sub)
        #     return code_max_count
        #
        # def plot_reportingset(self, rs, head, sheet, x, y):
        #     print(head)
        #     rs.print_content()
        #     x, y = (x, y)
        #     sheet.write(y, x, 'TOTAL')
        #     order_count = rs.split_main.keyOrder
        #     code_count = 0
        #     for code in order_count:
        #         try:
        #             code_label = rs.split_main[str(code)]
        #         except:
        #             try:
        #                 code_label = rs.split_main[int(code)]
        #             except:
        #                 code_label = rs.split_main[code]
        #
        #         code_label = code_label.decode('utf-8')
        #         sheet.write(y + 3 + code_count, x - 1, code_label)
        #         code_count += 1
        #
        #     code_offset = 0
        #     for label in self.reporting_columns:
        #         sheet.write(y + 2, x + code_offset, label.decode('utf-8'))
        #         code_count = 0
        #         for code in order_count:
        #             try:
        #                 sheet.write(y + 3 + code_count, x + code_offset, rs['TOTAL'][label][code])
        #             except:
        #                 try:
        #                     sheet.write(y + 3 + code_count, x + code_offset, rs['TOTAL'][label][str(code)])
        #                 except:
        #                     sheet.write(y + 3 + code_count, x + code_offset, 0)
        #             code_count += 1
        #         code_offset += 1
        #     code_max_count = code_count
        #
        #     x += len(self.reporting_columns)
        #     for split_sub_name in rs.subSplits:
        #         sheet.write(y, x, split_sub_name.decode('utf-8'))
        #         code_offset = 0
        #         splits_sub = rs.subSplits[split_sub_name]
        #         for code in splits_sub:
        #             try:
        #                 sheet.write(y + 1, x + code_offset, code.decode('utf-8'))
        #             except:
        #                 sheet.write(y + 1, x + code_offset, code)
        #             for label in self.reporting_columns:
        #                 sheet.write(y + 2, x + code_offset, label.decode('utf-8'))
        #                 # codeOffset += len(self.reporting_columns)
        #                 code_count = 0
        #                 for valCode in order_count:
        #                     try:
        #                         sheet.write(y + 3 + code_count, x + code_offset,
        #                                     rs[split_sub_name + "_" + str(code)][label][valCode])
        #                     except:
        #                         try:
        #                             sheet.write(y + 3 + code_count, x + code_offset,
        #                                         rs[split_sub_name + "_" + str(code)][label][str(valCode)])
        #                         except:
        #                             sheet.write(y + 3 + code_count, x + code_offset, 0)
        #                     code_count += 1
        #                 code_offset += 1
        #         x += len(self.reporting_columns) * len(splits_sub)
        #     return code_max_count
